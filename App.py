{
 "cells": [
  {
   "cell_type": "code",
   "execution_count": 4,
   "id": "6a038dc9-0d17-4401-97d1-c8bdc447e553",
   "metadata": {},
   "outputs": [
    {
     "name": "stderr",
     "output_type": "stream",
     "text": [
      "2024-07-09 12:19:12.762 \n",
      "  \u001b[33m\u001b[1mWarning:\u001b[0m to view this Streamlit app on a browser, run it with the following\n",
      "  command:\n",
      "\n",
      "    streamlit run C:\\Users\\chhi3001\\AppData\\Local\\Programs\\Python\\Python312\\Lib\\site-packages\\ipykernel_launcher.py [ARGUMENTS]\n",
      "2024-07-09 12:19:12.762 Session state does not function when running a script without `streamlit run`\n"
     ]
    }
   ],
   "source": [
    "import streamlit as st\n",
    "import pandas as pd\n",
    "from openpyxl import load_workbook\n",
    "from openpyxl import Workbook\n",
    "import shutil\n",
    "\n",
    "def process_files(input_folder):\n",
    "    # Loop through files in the directory\n",
    "    for filename in os.listdir(input_folder):\n",
    "        if filename.startswith('EXP_FV') and filename.endswith('.xlsx'):\n",
    "            full_path = os.path.join(input_folder, filename)\n",
    "            \n",
    "            # Load the Excel file\n",
    "            wb = load_workbook(full_path, data_only=True)\n",
    "            \n",
    "            # Process 'DUP Queries' sheet\n",
    "            if 'DUP Queries' in wb.sheetnames:\n",
    "                # Create a new workbook\n",
    "                new_wb = Workbook()\n",
    "                source_sheet = wb['DUP Queries']\n",
    "                new_sheet = new_wb.active\n",
    "                new_sheet.title = 'DUP_DAFeedback'\n",
    "                for row in source_sheet.iter_rows(values_only=True):\n",
    "                    new_sheet.append(row)\n",
    "                \n",
    "                # Save and clean dataframe\n",
    "                output_filename1 = 'DUP_DAFeedback.xlsx'\n",
    "                new_wb.save(output_filename1)\n",
    "                df1 = pd.read_excel(output_filename1)\n",
    "                df1['DA_FEEDBACK'] = df1['DA_FEEDBACK'].str.upper()\n",
    "                df_cleaned = df1[df1['DA_FEEDBACK'].notna() & (df1['DA_FEEDBACK'] != '')]\n",
    "                columns_to_keep = ['RESPONSE_ID', 'QUERY_ID', 'OLD_VALUE', 'DA_FEEDBACK']\n",
    "                df_filtered = df_cleaned[columns_to_keep]\n",
    "                df_filtered.to_excel(output_filename1, index=False)\n",
    "                \n",
    "                # Check feedback\n",
    "                column_to_check = 'DA_FEEDBACK'\n",
    "                other_values_present = ~df_filtered[column_to_check].isin(['OK', 'CHANGED']).any()\n",
    "                if other_values_present:\n",
    "                    st.info(\"'OK' or 'CHANGED' found, FILE IS OK\")\n",
    "                else:\n",
    "                    st.warning(\"Other feedback found, CHECK THE FILE\")\n",
    "                \n",
    "                # Move the file\n",
    "                output_path = os.path.join(input_folder, output_filename1)\n",
    "                shutil.move(output_filename1, output_path)\n",
    "            \n",
    "            # Process 'FV Feedback' sheet\n",
    "            if 'FV Feedback' in wb.sheetnames:\n",
    "                # Create a new workbook\n",
    "                new_wb = Workbook()\n",
    "                source_sheet = wb['FV Feedback']\n",
    "                new_sheet = new_wb.active\n",
    "                new_sheet.title = 'FV_DAFeedback'\n",
    "                for row in source_sheet.iter_rows(values_only=True):\n",
    "                    new_sheet.append(row)\n",
    "                \n",
    "                # Save and clean dataframe\n",
    "                output_filename2 = 'FV_DAFeedback.xlsx'\n",
    "                new_wb.save(output_filename2)\n",
    "                df2 = pd.read_excel(output_filename2)\n",
    "                df2['DA_FEEDBACK'] = df2['DA_FEEDBACK'].str.upper()\n",
    "                df_cleaned2 = df2[df2['DA_FEEDBACK'].notna() & (df2['DA_FEEDBACK'] != '')]\n",
    "                columns_to_keep = ['RESPONSE_ID', 'QUERY_ID', 'OLD_VALUE', 'DA_FEEDBACK']\n",
    "                df_filtered2 = df_cleaned2[columns_to_keep]\n",
    "                df_filtered2.to_excel(output_filename2, index=False)\n",
    "                \n",
    "                # Check feedback\n",
    "                column_to_check = 'DA_FEEDBACK'\n",
    "                other_values_present = ~df_filtered2[column_to_check].isin(['OK', 'CHANGED']).any()\n",
    "                if other_values_present:\n",
    "                    st.info(\"'OK' or 'CHANGED' found, FILE IS OK\")\n",
    "                else:\n",
    "                    st.warning(\"Other feedback found, CHECK THE FILE\")\n",
    "                \n",
    "                # Move the file\n",
    "                output_path = os.path.join(input_folder, output_filename2)\n",
    "                shutil.move(output_filename2, output_path)\n",
    "    \n",
    "    st.success(\"Files have been processed successfully.\")\n",
    "\n",
    "# Streamlit UI\n",
    "def main():\n",
    "    st.title(\"Get DA_Feedback Files\")\n",
    "    input_folder = st.text_input(\"Directory Path:\")\n",
    "    if st.button(\"Run\"):\n",
    "        process_files(input_folder)\n",
    "\n",
    "if __name__ == \"__main__\":\n",
    "    main()\n"
   ]
  },
  {
   "cell_type": "code",
   "execution_count": "null",
   "id": "6d28594b-8a4f-4caa-918b-a22c7d0ae7c4",
   "metadata": {},
   "outputs": [],
   "source": []
  }
 ],
 "metadata": {
  "kernelspec": {
   "display_name": "Python 3 (ipykernel)",
   "language": "python",
   "name": "python3"
  },
  "language_info": {
   "codemirror_mode": {
    "name": "ipython",
    "version": 3
   },
   "file_extension": ".py",
   "mimetype": "text/x-python",
   "name": "python",
   "nbconvert_exporter": "python",
   "pygments_lexer": "ipython3",
   "version": "3.12.3"
  }
 },
 "nbformat": 4,
 "nbformat_minor": 5
}
